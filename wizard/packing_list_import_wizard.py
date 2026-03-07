# -*- coding: utf-8 -*-
from odoo import models, fields, _
from odoo.exceptions import UserError
import base64
import io
import json
import logging
import re

_logger = logging.getLogger(__name__)


class _PLCellsIndex:
    """Clase para normalizar el acceso a celdas de Odoo Spreadsheet"""

    def __init__(self):
        self._cells = {}

    def put(self, col, row, content, source="unknown"):
        if col is None or row is None:
            return

        key = (int(col), int(row))
        if content in (None, False, ""):
            if key in self._cells:
                del self._cells[key]
        else:
            self._cells[key] = str(content)

    def ingest_cells(self, raw_cells):
        if not raw_cells:
            return

        for key, cell_data in raw_cells.items():
            col, row = self._parse_cell_key(key)
            if col is None or row is None:
                continue

            content = self._extract_content(cell_data)
            if content not in (None, False, ""):
                self.put(col, row, content, source="snapshot")

    def _parse_cell_key(self, key):
        if isinstance(key, str) and key and key[0].isalpha():
            match = re.match(r"^([A-Z]+)(\d+)$", key.upper())
            if match:
                col_str, row_str = match.groups()
                col = 0
                for char in col_str:
                    col = col * 26 + (ord(char) - ord("A") + 1)
                return col - 1, int(row_str) - 1

        if isinstance(key, str) and "," in key:
            parts = key.split(",")
            if len(parts) == 2:
                try:
                    return int(parts[0]), int(parts[1])
                except Exception:
                    return None, None

        return None, None

    def _extract_content(self, cell_data):
        if isinstance(cell_data, dict):
            return (
                cell_data.get("content")
                or cell_data.get("value")
                or cell_data.get("text")
                or ""
            )
        return cell_data or ""

    def apply_revision_commands(self, commands, target_sheet_id):
        applied = 0

        for cmd in commands:
            if isinstance(cmd, list):
                applied += self.apply_revision_commands(cmd, target_sheet_id)
                continue

            if not isinstance(cmd, dict):
                continue

            if cmd.get("sheetId") and cmd.get("sheetId") != target_sheet_id:
                continue

            cmd_type = cmd.get("type")

            if cmd_type == "UPDATE_CELL":
                col, row = cmd.get("col"), cmd.get("row")
                if col is not None and row is not None:
                    content = self._extract_content(cmd)
                    self.put(col, row, content, source="UPDATE_CELL_REV")
                    applied += 1

            elif cmd_type == "REMOVE_COLUMNS_ROWS":
                if cmd.get("dimension") == "row":
                    elements = sorted(cmd.get("elements", []), reverse=True)
                    for row_idx in elements:
                        self._shift_rows_up(row_idx)
                    applied += 1

            elif cmd_type in ("DELETE_CONTENT", "CLEAR_CELL"):
                zones = cmd.get("zones") or cmd.get("target") or []
                if isinstance(zones, dict):
                    zones = [zones]

                for zone in zones:
                    top = zone.get("top", 0)
                    bottom = zone.get("bottom", 0)
                    left = zone.get("left", 0)
                    right = zone.get("right", 0)

                    for r in range(top, bottom + 1):
                        for c in range(left, right + 1):
                            self.put(c, r, "", source="DELETE_REV")
                applied += 1

        return applied

    def _shift_rows_up(self, removed_row):
        new_cells = {}
        for (c, r), val in self._cells.items():
            if r < removed_row:
                new_cells[(c, r)] = val
            elif r > removed_row:
                new_cells[(c, r - 1)] = val
        self._cells = new_cells

    def value(self, col, row):
        return self._cells.get((int(col), int(row)))


class PackingListImportWizard(models.TransientModel):
    _name = "packing.list.import.wizard"
    _description = "Importar Packing List"

    picking_id = fields.Many2one(
        "stock.picking", string="Recepción", required=True, readonly=True
    )
    spreadsheet_id = fields.Many2one(
        "documents.document",
        related="picking_id.spreadsheet_id",
        readonly=True,
    )
    excel_file = fields.Binary(string="Archivo Excel", required=False, attachment=False)
    excel_filename = fields.Char(string="Nombre del archivo")

    def action_import_excel(self):
        self.ensure_one()
        _logger.info("=== [PL_IMPORT] INICIO PROCESO DE CARGA ===")

        rows = []
        if self.excel_file:
            _logger.info("[PL_IMPORT] Fuente seleccionada: archivo Excel")
            rows = self._get_data_from_excel_file()
        elif self.spreadsheet_id:
            _logger.info("[PL_IMPORT] Fuente seleccionada: spreadsheet")
            rows = self._get_data_from_spreadsheet()
        else:
            _logger.warning("[PL_IMPORT] No se recibió excel_file ni spreadsheet_id")

        _logger.info("[PL_IMPORT] Resultado Final: %s filas listas para importar.", len(rows))

        if not rows:
            raise UserError(
                _(
                    "No se encontraron datos válidos para importar. "
                    "Revise que la hoja contenga un producto reconocible y filas con cantidades/dimensiones mayores a cero."
                )
            )

        # --- LÓGICA DE LIMPIEZA PROFUNDA ---
        _logger.info("[PL_CLEANUP] Borrando datos previos...")
        old_move_lines = self.picking_id.move_line_ids
        old_lots = old_move_lines.mapped("lot_id")

        if old_move_lines:
            old_move_lines.write({"qty_done": 0})
            self.env.flush_all()

        if old_lots:
            quants = self.env["stock.quant"].sudo().search([("lot_id", "in", old_lots.ids)])
            if quants:
                quants.sudo().unlink()

        if old_move_lines:
            old_move_lines.unlink()

        for lot in old_lots:
            if self.env["stock.move.line"].search_count([("lot_id", "=", lot.id)]) == 0:
                try:
                    with self.env.cr.savepoint():
                        lot.unlink()
                except Exception as e:
                    _logger.warning("[PL_CLEANUP] No se pudo borrar lote %s: %s", lot.name, e)

        # --- CREACIÓN DE NUEVOS REGISTROS ---
        move_lines_created = 0
        skipped_without_move = 0
        skipped_qty_zero = 0
        next_prefix = self._get_next_global_prefix()
        containers = {}

        for data in rows:
            product = data["product"]
            move = self.picking_id.move_ids.filtered(lambda m: m.product_id == product)[:1]

            if not move:
                skipped_without_move += 1
                _logger.warning(
                    "[PL_IMPORT] No existe stock.move para producto '%s' en picking %s. Fila omitida.",
                    product.display_name,
                    self.picking_id.name,
                )
                continue

            unit_type = data.get("tipo", "Placa")

            qty_done = 0.0
            final_alto = 0.0
            final_ancho = 0.0

            if unit_type == "Placa":
                final_alto = data.get("alto", 0.0)
                final_ancho = data.get("ancho", 0.0)
                qty_done = round(final_alto * final_ancho, 3)
            else:
                qty_done = data.get("quantity", 0.0)
                final_alto = 0.0
                final_ancho = 0.0

            if qty_done <= 0:
                skipped_qty_zero += 1
                _logger.info(
                    "[PL_IMPORT] Fila omitida por qty_done<=0 | product=%s | tipo=%s | alto=%s | ancho=%s | quantity=%s",
                    product.display_name,
                    unit_type,
                    data.get("alto"),
                    data.get("ancho"),
                    data.get("quantity"),
                )
                continue

            cont = (data.get("contenedor") or "SN").strip() or "SN"

            if cont not in containers:
                containers[cont] = {
                    "pre": str(next_prefix),
                    "num": self._get_next_lot_number_for_prefix(str(next_prefix)),
                }
                next_prefix += 1

            l_name = f"{containers[cont]['pre']}-{containers[cont]['num']:02d}"

            grupo_ids = []
            if data.get("grupo_name"):
                grupo_name = data["grupo_name"].strip()
                grupo = self.env["stock.lot.group"].search([("name", "=", grupo_name)], limit=1)
                if not grupo:
                    grupo = self.env["stock.lot.group"].create({"name": grupo_name})
                grupo_ids = [grupo.id]

            lot_selection_value = str(unit_type).lower()

            lot = self.env["stock.lot"].create({
                "name": l_name,
                "product_id": product.id,
                "company_id": self.picking_id.company_id.id,
                "x_grosor": data.get("grosor"),
                "x_alto": final_alto,
                "x_ancho": final_ancho,
                "x_color": data.get("color"),
                "x_bloque": data.get("bloque"),
                "x_numero_placa": data.get("numero_placa"),
                "x_atado": data.get("atado"),
                "x_tipo": lot_selection_value,
                "x_grupo": [(6, 0, grupo_ids)],
                "x_pedimento": data.get("pedimento"),
                "x_contenedor": cont,
                "x_referencia_proveedor": data.get("ref_proveedor"),
            })

            self.env["stock.move.line"].create({
                "move_id": move.id,
                "product_id": product.id,
                "lot_id": lot.id,
                "qty_done": qty_done,
                "location_id": self.picking_id.location_id.id,
                "location_dest_id": self.picking_id.location_dest_id.id,
                "picking_id": self.picking_id.id,
                "x_grosor_temp": data.get("grosor"),
                "x_alto_temp": final_alto,
                "x_ancho_temp": final_ancho,
                "x_color_temp": data.get("color"),
                "x_tipo_temp": lot_selection_value,
                "x_bloque_temp": data.get("bloque"),
                "x_atado_temp": data.get("atado"),
                "x_pedimento_temp": data.get("pedimento"),
                "x_contenedor_temp": cont,
                "x_referencia_proveedor_temp": data.get("ref_proveedor"),
                "x_grupo_temp": [(6, 0, grupo_ids)],
            })

            containers[cont]["num"] += 1
            move_lines_created += 1

        # --- SINCRONIZACIÓN WORKSHEET ---
        if self.picking_id.ws_spreadsheet_id:
            try:
                self.picking_id.ws_spreadsheet_id.sudo().unlink()
                self.picking_id.write({"worksheet_imported": False})
                _logger.info("[PL_IMPORT] Worksheet antiguo eliminado para forzar sincronización.")
            except Exception as e:
                _logger.warning("[PL_IMPORT] No se pudo eliminar el Worksheet anterior: %s", e)

        self.picking_id.write({"packing_list_imported": True})

        # ── SINCRONIZAR CANTIDADES EN LÍNEAS DE LA OC ─────────────────────────
        self._sync_quantities_to_po_lines()
        # ──────────────────────────────────────────────────────────────────────

        _logger.info(
            "=== [PL_IMPORT] PROCESO TERMINADO. Creados %s registros | sin move: %s | qty=0: %s ===",
            move_lines_created,
            skipped_without_move,
            skipped_qty_zero,
        )

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "PL Procesado",
                "message": (
                    f"Se han importado/corregido {move_lines_created} lotes. "
                    f"Omitidos sin movimiento: {skipped_without_move}. "
                    f"Omitidos por cantidad 0: {skipped_qty_zero}. "
                    "El Worksheet ha sido reiniciado."
                ),
                "type": "success",
                "next": {"type": "ir.actions.act_window_close"},
            },
        }

    def _sync_quantities_to_po_lines(self):
        picking = self.picking_id
        po = self.env["purchase.order"].search([("picking_ids", "in", picking.id)], limit=1)

        if not po:
            _logger.warning("[PL_SYNC] No se encontró PO asociada al picking.")
            return

        for po_line in po.order_line:
            product = po_line.product_id
            move_lines = picking.move_line_ids.filtered(lambda ml: ml.product_id == product)
            total_embarcado = sum(move_lines.mapped("qty_done"))

            if total_embarcado <= 0:
                continue

            vals = {"x_qty_embarcada": total_embarcado}
            if not po_line.x_qty_solicitada_original:
                vals["x_qty_solicitada_original"] = po_line.product_qty
            vals["product_qty"] = total_embarcado
            po_line.write(vals)

        _logger.info("[PL_SYNC] Cantidades sincronizadas a la OC %s.", po.name)

    def _get_data_from_spreadsheet(self):
        doc = self.spreadsheet_id
        if not doc:
            _logger.warning("[PL_DEBUG] No hay spreadsheet_id relacionado al picking")
            return []

        _logger.info(
            "[PL_DEBUG] Doc ID: %s | snapshot: %s | data: %s",
            doc.id,
            bool(doc.spreadsheet_snapshot),
            bool(doc.spreadsheet_data),
        )

        spreadsheet_json = self._get_current_spreadsheet_state(doc)
        if not spreadsheet_json or not spreadsheet_json.get("sheets"):
            _logger.warning("[PL_DEBUG] spreadsheet_json vacío o sin sheets")
            return []

        _logger.info(
            "[PL_DEBUG] Sheets encontrados: %s",
            [s.get("name") for s in spreadsheet_json.get("sheets", [])],
        )

        all_rows = []
        products_not_found = []

        for sheet in spreadsheet_json["sheets"]:
            idx = _PLCellsIndex()
            idx.ingest_cells(sheet.get("cells", {}))
            _logger.info(
                "[PL_DEBUG] Sheet '%s': %s celdas tras ingest",
                sheet.get("name"),
                len(idx._cells),
            )

            product = self._identify_product_from_sheet(idx)
            _logger.info(
                "[PL_DEBUG] Producto identificado en hoja '%s': %s",
                sheet.get("name"),
                product.name if product else "NINGUNO",
            )

            if not product:
                products_not_found.append(sheet.get("name"))
                continue

            sheet_rows = self._extract_rows_from_index(idx, product)
            _logger.info(
                "[PL_DEBUG] Filas extraídas para '%s': %s",
                product.name,
                len(sheet_rows),
            )
            all_rows.extend(sheet_rows)

        if products_not_found:
            _logger.warning(
                "[PL_DEBUG] Hojas sin producto identificado: %s",
                products_not_found,
            )

        return all_rows

    def _get_current_spreadsheet_state(self, doc):
        snapshot_len = len(doc.spreadsheet_snapshot) if doc.spreadsheet_snapshot else 0
        data_len = len(doc.spreadsheet_data) if doc.spreadsheet_data else 0

        _logger.info(
            "[PL_DEBUG] snapshot existe: %s | len: %s",
            bool(doc.spreadsheet_snapshot),
            snapshot_len,
        )
        _logger.info(
            "[PL_DEBUG] spreadsheet_data existe: %s | len: %s",
            bool(doc.spreadsheet_data),
            data_len,
        )

        if doc.spreadsheet_snapshot:
            try:
                parsed = self._safe_json_load(doc.spreadsheet_snapshot)
                if parsed:
                    sheets_count = len(parsed.get("sheets", []))
                    _logger.info(
                        "[PL_DEBUG] snapshot parseado OK | sheets: %s | revisionId: %s",
                        sheets_count,
                        parsed.get("revisionId", "N/A"),
                    )
                    if parsed.get("sheets"):
                        return self._apply_pending_revisions(doc, parsed)
            except Exception as e:
                _logger.warning("[PL_IMPORT] Error leyendo snapshot: %s", e)

        try:
            if hasattr(doc, "_get_spreadsheet_serialized_snapshot"):
                snapshot_data = doc._get_spreadsheet_serialized_snapshot()
                _logger.info(
                    "[PL_DEBUG] _get_spreadsheet_serialized_snapshot: %s",
                    bool(snapshot_data),
                )
                if snapshot_data:
                    parsed = self._safe_json_load(snapshot_data)
                    if parsed and parsed.get("sheets"):
                        return self._apply_pending_revisions(doc, parsed)
        except Exception as e:
            _logger.warning("[PL_IMPORT] Error en _get_spreadsheet_serialized_snapshot: %s", e)

        _logger.info("[PL_IMPORT] Fallback: spreadsheet_data + todas las revisiones")
        return self._load_spreadsheet_with_all_revisions(doc)

    def _apply_pending_revisions(self, doc, spreadsheet_json):
        snapshot_revision_id = spreadsheet_json.get("revisionId", "")
        _logger.info(
            "[PL_DEBUG] _apply_pending_revisions | revisionId snapshot: '%s'",
            snapshot_revision_id,
        )

        if not snapshot_revision_id:
            _logger.info("[PL_DEBUG] Sin revisionId en snapshot, retornando json tal cual")
            return spreadsheet_json

        revisions = self.env["spreadsheet.revision"].sudo().with_context(active_test=False).search([
            ("res_model", "=", "documents.document"),
            ("res_id", "=", doc.id),
        ], order="id asc")
        _logger.info("[PL_DEBUG] Revisiones totales en BD: %s", len(revisions))

        start_applying = False
        all_cmds = []

        for rev in revisions:
            rev_data = self._safe_json_load(rev.commands)
            if not rev_data:
                continue

            if not start_applying:
                rev_id = rev_data.get("id") if isinstance(rev_data, dict) else None
                if rev_id == snapshot_revision_id:
                    start_applying = True
                continue

            if isinstance(rev_data, dict) and rev_data.get("type") == "SNAPSHOT_CREATED":
                continue

            if isinstance(rev_data, dict) and "commands" in rev_data:
                all_cmds.extend(rev_data["commands"])
            elif isinstance(rev_data, list):
                all_cmds.extend(rev_data)

        _logger.info(
            "[PL_DEBUG] Comandos pendientes a aplicar tras snapshot: %s",
            len(all_cmds),
        )

        if not all_cmds:
            _logger.info("[PL_DEBUG] Sin comandos pendientes, retornando snapshot directo")
            return spreadsheet_json

        for sheet in spreadsheet_json.get("sheets", []):
            sheet_id = sheet.get("id")
            idx = _PLCellsIndex()
            cells_before = len(sheet.get("cells", {}))
            idx.ingest_cells(sheet.get("cells", {}))
            applied = idx.apply_revision_commands(all_cmds, sheet_id)

            _logger.info(
                "[PL_DEBUG] Sheet '%s' | celdas antes: %s | cmds aplicados: %s | celdas después: %s",
                sheet.get("name"),
                cells_before,
                applied,
                len(idx._cells),
            )

            sheet["cells"] = {
                f"{self._col_to_letter(c)}{r + 1}": {"content": v}
                for (c, r), v in idx._cells.items()
            }

        return spreadsheet_json

    def _load_spreadsheet_with_all_revisions(self, doc):
        spreadsheet_json = self._load_spreadsheet_json(doc)
        _logger.info(
            "[PL_DEBUG] _load_spreadsheet_json: %s | sheets: %s",
            bool(spreadsheet_json),
            len(spreadsheet_json.get("sheets", [])) if spreadsheet_json else 0,
        )

        if not spreadsheet_json:
            return None

        revisions = self.env["spreadsheet.revision"].sudo().with_context(active_test=False).search([
            ("res_model", "=", "documents.document"),
            ("res_id", "=", doc.id),
        ], order="id asc")
        _logger.info("[PL_DEBUG] Revisiones en fallback: %s", len(revisions))

        all_cmds = []
        for rev in revisions:
            rev_data = self._safe_json_load(rev.commands)
            if not rev_data:
                continue

            if isinstance(rev_data, dict) and rev_data.get("type") == "SNAPSHOT_CREATED":
                continue

            if isinstance(rev_data, dict) and "commands" in rev_data:
                all_cmds.extend(rev_data["commands"])
            elif isinstance(rev_data, list):
                all_cmds.extend(rev_data)

        _logger.info("[PL_DEBUG] Comandos totales en fallback: %s", len(all_cmds))

        for sheet in spreadsheet_json.get("sheets", []):
            idx = _PLCellsIndex()
            cells_before = len(sheet.get("cells", {}))
            idx.ingest_cells(sheet.get("cells", {}))
            applied = idx.apply_revision_commands(all_cmds, sheet.get("id"))

            _logger.info(
                "[PL_DEBUG] Fallback sheet '%s' | celdas antes: %s | cmds: %s | celdas después: %s",
                sheet.get("name"),
                cells_before,
                applied,
                len(idx._cells),
            )

            sample = list(idx._cells.items())[:10]
            _logger.info("[PL_DEBUG] Muestra celdas: %s", sample)

            sheet["cells"] = {
                f"{self._col_to_letter(c)}{r + 1}": {"content": v}
                for (c, r), v in idx._cells.items()
            }

        return spreadsheet_json

    def _safe_json_load(self, payload):
        if not payload:
            return None

        try:
            if isinstance(payload, bytes):
                payload = payload.decode("utf-8")
            if isinstance(payload, str):
                payload = payload.strip()
                if not payload:
                    return None
                return json.loads(payload)
            if isinstance(payload, dict):
                return payload
            if isinstance(payload, list):
                return payload
        except Exception as e:
            _logger.warning("[PL_DEBUG] No se pudo parsear JSON: %s", e)

        return None

    def _col_to_letter(self, col):
        result = ""
        col += 1
        while col:
            col, remainder = divmod(col - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _normalize_product_text(self, text):
        if not text:
            return ""

        text = str(text).strip()
        text = re.sub(r"\(\s*\)", "", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def _extract_short_product_name(self, text):
        if not text:
            return ""

        short_name = str(text).split("(")[0].strip()
        short_name = re.sub(r"\s+", " ", short_name)
        return short_name.strip()

    def _find_product_by_header(self, raw_product_value):
        Product = self.env["product.product"]

        raw_name = str(raw_product_value or "").strip()
        clean_name = self._normalize_product_text(raw_name)
        short_name = self._extract_short_product_name(clean_name)

        _logger.info(
            "[PL_DEBUG] Buscando producto | raw='%s' | clean='%s' | short='%s'",
            raw_name,
            clean_name,
            short_name,
        )

        search_attempts = []

        if clean_name:
            search_attempts.extend([
                ("name", "=", clean_name, "name exacto clean"),
                ("default_code", "=", clean_name, "default_code exacto clean"),
                ("name", "ilike", clean_name, "name ilike clean"),
                ("default_code", "ilike", clean_name, "default_code ilike clean"),
            ])

        if short_name and short_name != clean_name:
            search_attempts.extend([
                ("name", "=", short_name, "name exacto short"),
                ("default_code", "=", short_name, "default_code exacto short"),
                ("name", "ilike", short_name, "name ilike short"),
                ("default_code", "ilike", short_name, "default_code ilike short"),
            ])

        for field_name, operator, value, label in search_attempts:
            product = Product.search([(field_name, operator, value)], limit=1)
            if product:
                _logger.info(
                    "[PL_DEBUG] Producto encontrado por %s: %s",
                    label,
                    product.display_name,
                )
                return product

        _logger.warning(
            "[PL_DEBUG] No se encontró producto para raw='%s' | clean='%s' | short='%s'",
            raw_name,
            clean_name,
            short_name,
        )
        return None

    def _identify_product_from_sheet(self, idx):
        p_info = None

        for r in range(3):
            label = str(idx.value(0, r) or "").upper().strip()
            val_b = idx.value(1, r)
            _logger.info("[PL_DEBUG] identify fila %s: A='%s' B='%s'", r, label, val_b)
            if "PRODUCTO:" in label:
                p_info = val_b
                break

        if not p_info:
            p_info = idx.value(1, 0)

        if not p_info:
            _logger.warning("[PL_DEBUG] No se encontró info de producto en la hoja")
            return None

        return self._find_product_by_header(p_info)

    def _extract_rows_from_index(self, idx, product):
        rows = []
        unit_type = product.product_tmpl_id.x_unidad_del_producto or "Placa"
        _logger.info(
            "[PL_DEBUG] Extrayendo filas para '%s' | unit_type: %s",
            product.name,
            unit_type,
        )

        if unit_type == "Placa":
            idx_notas = 4
            idx_bloque = 5
            idx_placa = 6
            idx_atado = 7
            idx_grupo = 8
            idx_pedimento = 9
            idx_contenedor = 10
            idx_ref = 11
        else:
            idx_notas = 3
            idx_bloque = 4
            idx_placa = 5
            idx_atado = 6
            idx_grupo = 7
            idx_pedimento = 8
            idx_contenedor = 9
            idx_ref = 10

        filas_validas = 0
        filas_invalidas = 0

        for r in range(3, 300):
            raw_a = idx.value(0, r)
            raw_b = idx.value(1, r)
            raw_c = idx.value(2, r)

            val_b = self._to_float(raw_b)
            val_c = self._to_float(raw_c)

            es_valido = False
            if unit_type == "Placa":
                if val_b > 0 and val_c > 0:
                    es_valido = True
            else:
                if val_b > 0:
                    es_valido = True

            if es_valido:
                filas_validas += 1
                rows.append({
                    "product": product,
                    "grosor": str(raw_a or "").strip(),
                    "alto": val_b if unit_type == "Placa" else 0.0,
                    "ancho": val_c if unit_type == "Placa" else 0.0,
                    "quantity": val_b if unit_type != "Placa" else 0.0,
                    "color": str(idx.value(idx_notas, r) or "").strip(),
                    "bloque": str(idx.value(idx_bloque, r) or "").strip(),
                    "numero_placa": str(idx.value(idx_placa, r) or "").strip(),
                    "atado": str(idx.value(idx_atado, r) or "").strip(),
                    "tipo": unit_type,
                    "grupo_name": str(idx.value(idx_grupo, r) or "").strip(),
                    "pedimento": str(idx.value(idx_pedimento, r) or "").strip(),
                    "contenedor": str(idx.value(idx_contenedor, r) or "SN").strip(),
                    "ref_proveedor": str(idx.value(idx_ref, r) or "").strip(),
                })
            else:
                if filas_invalidas < 5 and (raw_a or raw_b or raw_c):
                    _logger.info(
                        "[PL_DEBUG] Fila %s inválida | A='%s' B='%s' C='%s'",
                        r + 1,
                        raw_a,
                        raw_b,
                        raw_c,
                    )
                    filas_invalidas += 1

        _logger.info(
            "[PL_DEBUG] Total filas válidas: %s | inválidas con contenido: %s",
            filas_validas,
            filas_invalidas,
        )
        return rows

    def _to_float(self, val):
        if val in (None, False, ""):
            return 0.0

        try:
            txt = str(val).strip()
            txt = txt.replace(" ", "")
            txt = txt.replace(",", ".")
            return float(txt)
        except Exception:
            return 0.0

    def _get_next_global_prefix(self):
        self.env.cr.execute(
            """
            SELECT CAST(SUBSTRING(name FROM '^([0-9]+)-') AS INTEGER) as prefix_num
            FROM stock_lot
            WHERE name ~ '^[0-9]+-[0-9]+$'
              AND company_id = %s
            ORDER BY prefix_num DESC
            LIMIT 1
            """,
            (self.picking_id.company_id.id,),
        )
        res = self.env.cr.fetchone()
        return (res[0] + 1) if res and res[0] else 1

    def _get_next_lot_number_for_prefix(self, prefix):
        self.env.cr.execute(
            """
            SELECT name
            FROM stock_lot
            WHERE name LIKE %s
              AND company_id = %s
            ORDER BY CAST(SUBSTRING(name FROM '-([0-9]+)$') AS INTEGER) DESC
            LIMIT 1
            """,
            (f"{prefix}-%", self.picking_id.company_id.id),
        )
        res = self.env.cr.fetchone()
        return int(res[0].split("-")[1]) + 1 if res else 1

    def _load_spreadsheet_json(self, doc):
        if not doc.spreadsheet_data:
            return None

        try:
            return self._safe_json_load(doc.spreadsheet_data)
        except Exception as e:
            _logger.warning("[PL_DEBUG] Error leyendo spreadsheet_data: %s", e)
            return None

    def _get_data_from_excel_file(self):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(base64.b64decode(self.excel_file)), data_only=True)
        rows = []

        for sheet in wb.worksheets:
            p_info = sheet["B1"].value
            if not p_info:
                _logger.warning("[PL_DEBUG][XLSX] Hoja '%s' sin encabezado de producto en B1", sheet.title)
                continue

            product = self._find_product_by_header(p_info)
            if not product:
                _logger.warning(
                    "[PL_DEBUG][XLSX] No se encontró producto para hoja '%s' con encabezado '%s'",
                    sheet.title,
                    p_info,
                )
                continue

            unit_type = product.product_tmpl_id.x_unidad_del_producto or "Placa"

            if unit_type == "Placa":
                col_notas = 5
                col_bloque = 6
                col_placa = 7
                col_atado = 8
                col_grupo = 9
                col_pedimento = 10
                col_contenedor = 11
                col_ref = 12
            else:
                col_notas = 4
                col_bloque = 5
                col_placa = 6
                col_atado = 7
                col_grupo = 8
                col_pedimento = 9
                col_contenedor = 10
                col_ref = 11

            for r in range(4, sheet.max_row + 1):
                raw_b = sheet.cell(r, 2).value
                raw_c = sheet.cell(r, 3).value

                val_b = self._to_float(raw_b)
                val_c = self._to_float(raw_c)

                es_valido = False
                if unit_type == "Placa":
                    if val_b > 0 and val_c > 0:
                        es_valido = True
                else:
                    if val_b > 0:
                        es_valido = True

                if es_valido:
                    rows.append({
                        "product": product,
                        "grosor": str(sheet.cell(r, 1).value or "").strip(),
                        "alto": val_b if unit_type == "Placa" else 0.0,
                        "ancho": val_c if unit_type == "Placa" else 0.0,
                        "quantity": val_b if unit_type != "Placa" else 0.0,
                        "color": str(sheet.cell(r, col_notas).value or "").strip(),
                        "bloque": str(sheet.cell(r, col_bloque).value or "").strip(),
                        "numero_placa": str(sheet.cell(r, col_placa).value or "").strip(),
                        "atado": str(sheet.cell(r, col_atado).value or "").strip(),
                        "tipo": unit_type,
                        "grupo_name": str(sheet.cell(r, col_grupo).value or "").strip(),
                        "pedimento": str(sheet.cell(r, col_pedimento).value or "").strip(),
                        "contenedor": str(sheet.cell(r, col_contenedor).value or "SN").strip(),
                        "ref_proveedor": str(sheet.cell(r, col_ref).value or "").strip(),
                    })

        _logger.info("[PL_DEBUG][XLSX] Total filas extraídas desde Excel: %s", len(rows))
        return rows