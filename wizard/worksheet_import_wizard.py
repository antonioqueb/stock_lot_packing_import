# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import json
import logging

_logger = logging.getLogger(__name__)

class WorksheetImportWizard(models.TransientModel):
    _name = 'worksheet.import.wizard'
    _description = 'Importar Worksheet (Spreadsheet WS o Excel)'
    
    picking_id = fields.Many2one('stock.picking', string='Recepción', required=True, readonly=True)
    ws_spreadsheet_id = fields.Many2one('documents.document', related='picking_id.ws_spreadsheet_id', readonly=True)
    excel_file = fields.Binary(string='Archivo Excel (Opcional)', attachment=False)
    excel_filename = fields.Char(string='Nombre del archivo')
    state = fields.Selection([
        ('draft', 'Captura'),
        ('review', 'Resumen previo'),
    ], default='draft')
    summary_html = fields.Html(string='Resumen', readonly=True, sanitize=False)

    # =========================================================================
    # PASO 1: RESUMEN PREVIO (no toca nada)
    # =========================================================================

    def _ws_collect_rows(self):
        """Lee las filas del WS (spreadsheet o Excel) sin aplicar cambios."""
        self.ensure_one()

        if self.picking_id.state == 'done':
            raise UserError(
                'La recepción ya está validada. Las correcciones ahora solo '
                'proceden por ajuste manual de inventario.')

        if not self.ws_spreadsheet_id and not self.excel_file:
            raise UserError('No se encontró el Spreadsheet del Worksheet ni se subió un archivo Excel.')

        rows_data = self._get_data_from_excel() if self.excel_file else self._get_data_from_spreadsheet()
        if not rows_data:
            raise UserError('No se encontraron datos de medidas reales (Alto/Largo Real) para procesar.')
        return rows_data

    def _ws_find_move_line(self, product, lot_name):
        domain_base = [
            ('picking_id', '=', self.picking_id.id),
            ('lot_id.name', '=', lot_name),
        ]
        ml = self.env['stock.move.line'].search(
            domain_base + [('product_id', '=', product.id)], limit=1)
        if not ml:
            ml = self.env['stock.move.line'].search(domain_base, limit=1)
        return ml

    def _ws_build_summary_html(self, rows_data):
        picking = self.picking_id

        def _tipo(product):
            u = product.product_tmpl_id.x_unidad_del_producto or 'Placa'
            u = str(u).strip().lower()
            return {'placa': 'Placas', 'formato': 'Formatos'}.get(u, 'Piezas / Adhesivos')

        def _row_capture(d):
            if d.get('is_placa', True):
                return bool(d.get('alto_real') or d.get('ancho_real'))
            return bool(d.get('qty_real'))

        buckets = {}
        diffs = []
        not_found = []
        total_m2_real = 0.0
        total_units_real = 0.0

        for data in rows_data:
            product = data['product']
            lot_name = data['lot_name']
            ml = self._ws_find_move_line(product, lot_name)
            if not ml or not ml.lot_id:
                not_found.append(lot_name)
                continue

            tipo = _tipo(product)
            b = buckets.setdefault(tipo, {
                'captured': 0, 'missing': 0,
                'declared': 0.0, 'real': 0.0, 'missing_qty': 0.0,
            })

            declared = picking._ws_move_line_qty(ml)
            is_m2 = 'm²' in (ml.product_uom_id.name or '') or 'm2' in (ml.product_uom_id.name or '').lower()

            if not _row_capture(data):
                b['missing'] += 1
                b['missing_qty'] += declared
                continue

            if data.get('is_placa', True):
                real = round((data.get('alto_real') or 0.0) * (data.get('ancho_real') or 0.0), 3)
            else:
                real = data.get('qty_real') or 0.0

            b['captured'] += 1
            b['declared'] += declared
            b['real'] += real
            if is_m2:
                total_m2_real += real
            else:
                total_units_real += real

            if abs(real - declared) > 0.005:
                diffs.append((lot_name, product.display_name, declared, real))

        # ── HTML ──
        rows_html = ''
        tot_missing = 0
        for tipo in ('Placas', 'Formatos', 'Piezas / Adhesivos'):
            b = buckets.get(tipo)
            if not b:
                continue
            tot_missing += b['missing']
            delta = b['real'] - b['declared']
            color = '#dc3545' if abs(delta) > 0.005 or b['missing'] else '#198754'
            rows_html += (
                '<tr><td><b>%s</b></td><td class="text-end">%s</td>'
                '<td class="text-end">%s</td><td class="text-end">%.3f</td>'
                '<td class="text-end">%.3f</td>'
                '<td class="text-end" style="color:%s;font-weight:600;">%+.3f</td></tr>'
            ) % (tipo, b['captured'], b['missing'], b['declared'], b['real'], color, delta)

        html = (
            '<h5>Resumen previo — lo que se va a procesar</h5>'
            '<table class="table table-sm"><thead><tr>'
            '<th>Tipo</th><th class="text-end">Capturados</th>'
            '<th class="text-end">Faltantes</th>'
            '<th class="text-end">Declarado prov.</th>'
            '<th class="text-end">Real WS</th><th class="text-end">Dif.</th>'
            '</tr></thead><tbody>%s</tbody></table>'
            '<p><b>Totales reales:</b> %.3f m² · %.2f unidades</p>'
        ) % (rows_html, total_m2_real, total_units_real)

        if tot_missing:
            html += (
                '<div class="alert alert-danger"><b>⚠ %s lote(s) SIN captura</b>: '
                'al confirmar se eliminarán como faltantes. Si en realidad sí '
                'llegaron, captura sus medidas antes de confirmar.</div>'
            ) % tot_missing

        if diffs:
            d_rows = ''.join(
                '<tr><td>%s</td><td>%s</td><td class="text-end">%.3f</td>'
                '<td class="text-end">%.3f</td><td class="text-end" style="color:#dc3545;">%+.3f</td></tr>'
                % (l, p, dec, real, real - dec)
                for l, p, dec, real in diffs[:20]
            )
            more = '' if len(diffs) <= 20 else '<p class="text-muted">…y %s diferencias más.</p>' % (len(diffs) - 20)
            html += (
                '<h6>Diferencias contra lo declarado (%s)</h6>'
                '<table class="table table-sm"><thead><tr><th>Lote</th><th>Producto</th>'
                '<th class="text-end">Declarado</th><th class="text-end">Real</th>'
                '<th class="text-end">Dif.</th></tr></thead><tbody>%s</tbody></table>%s'
            ) % (len(diffs), d_rows, more)

        if not_found:
            html += (
                '<div class="alert alert-warning">Lotes del WS no encontrados en '
                'la recepción (se ignorarán): %s</div>'
            ) % ', '.join(not_found[:15])

        html += (
            '<div class="alert alert-info" style="margin-bottom:0;">'
            '<b>Reglas de corrección:</b> piezas/placas recibidas → Packing List. '
            'Medidas/cantidades reales → Worksheet (este paso). '
            'Después de validar la recepción → solo ajuste manual de inventario.</div>'
        )
        return html

    def _ws_reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'worksheet.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_review_worksheet(self):
        self.ensure_one()
        rows_data = self._ws_collect_rows()
        self.summary_html = self._ws_build_summary_html(rows_data)
        self.state = 'review'
        return self._ws_reopen()

    def action_back_to_draft(self):
        self.ensure_one()
        self.state = 'draft'
        return self._ws_reopen()

    def action_edit_worksheet(self):
        """Abre el WS para corregir capturas antes de confirmar."""
        self.ensure_one()
        return self.picking_id.action_open_worksheet_spreadsheet()
    
    def action_import_worksheet(self):
        self.ensure_one()

        rows_data = self._ws_collect_rows()

        # RED DE SEGURIDAD: si ninguna fila trae captura (todo vacío o en 0),
        # abortar sin tocar nada. Antes esto borraba TODOS los lotes de la
        # recepción al interpretarlos como faltantes.
        def _row_has_capture(d):
            if d.get('is_placa', True):
                return bool(d.get('alto_real') or d.get('ancho_real'))
            return bool(d.get('qty_real'))

        captured = [d for d in rows_data if _row_has_capture(d)]
        _logger.info(
            "[WS_IMPORT] Filas leídas: %s | con captura: %s",
            len(rows_data), len(captured),
        )
        if not captured:
            raise UserError(
                'No se encontró ninguna captura en el Worksheet: todas las '
                'filas tienen la cantidad o medida real vacía o en 0.\n\n'
                'No se modificó nada. Captura los valores reales en el WS, '
                'espera unos segundos a que la hoja guarde los cambios (o '
                'ciérrala) y vuelve a dar Procesar WS.'
            )

        lines_updated = 0
        total_missing_pieces = 0
        total_missing_m2 = 0
        
        container_lots = {}
        lots_to_delete = []
        move_lines_to_delete = []

        for data in rows_data:
            product = data['product']
            lot_name = data['lot_name']

            domain_base = [
                ('picking_id', '=', self.picking_id.id),
                ('lot_id.name', '=', lot_name)
            ]
            
            move_line = self.env['stock.move.line'].search(domain_base + [('product_id', '=', product.id)], limit=1)

            if not move_line:
                _logger.info(f"Fallback búsqueda lote: '{lot_name}' sin filtro de producto.")
                move_line = self.env['stock.move.line'].search(domain_base, limit=1)

            if not move_line or not move_line.lot_id:
                _logger.warning(f"No se encontró el lote '{lot_name}' para el producto {product.name} en esta recepción (Picking ID: {self.picking_id.id}).")
                continue

            lot = move_line.lot_id

            # Formatos: el WS solo corrobora cantidad real contra teórica.
            # No hay alto/largo real y no se renumeran lotes por contenedor.
            if not data.get('is_placa', True):
                qty_real = data.get('qty_real') or 0.0
                if qty_real == 0.0:
                    total_missing_pieces += 1
                    total_missing_m2 += self.picking_id._ws_move_line_qty(move_line)
                    move_lines_to_delete.append(move_line)
                    lots_to_delete.append(lot)
                else:
                    move_line.write({'qty_done': qty_real})
                    lines_updated += 1
                continue

            alto_real = data['alto_real']
            ancho_real = data['ancho_real']

            if alto_real == 0.0 and ancho_real == 0.0:
                m2_faltante = lot.x_alto * lot.x_ancho if lot.x_alto and lot.x_ancho else 0
                total_missing_pieces += 1
                total_missing_m2 += m2_faltante
                
                move_lines_to_delete.append(move_line)
                lots_to_delete.append(lot)
            else:
                lot.write({
                    'x_alto': alto_real,
                    'x_ancho': ancho_real
                })
                new_qty = round(alto_real * ancho_real, 3)
                move_line.write({
                    'qty_done': new_qty,
                    'x_alto_temp': alto_real,
                    'x_ancho_temp': ancho_real,
                })
                
                cont = lot.x_contenedor or 'SN'
                if cont not in container_lots:
                    container_lots[cont] = []
                container_lots[cont].append({
                    'lot': lot,
                    'original_name': lot.name,
                    'move_line': move_line
                })
                lines_updated += 1

        for ml in move_lines_to_delete:
            ml.write({'qty_done': 0})
        
        for lot in lots_to_delete:
            quants = self.env['stock.quant'].sudo().search([('lot_id', '=', lot.id)])
            if quants:
                quants.sudo().write({'quantity': 0, 'reserved_quantity': 0})
                quants.sudo().unlink()
        
        for ml in move_lines_to_delete:
            ml.unlink()
        
        for lot in lots_to_delete:
            other_ops = self.env['stock.move.line'].search([('lot_id', '=', lot.id)])
            if not other_ops:
                remaining_quants = self.env['stock.quant'].sudo().search([('lot_id', '=', lot.id)])
                if remaining_quants:
                    remaining_quants.sudo().unlink()
                lot.unlink()

        for cont, lot_data_list in container_lots.items():
            if not lot_data_list:
                continue
            
            lot_data_list.sort(key=lambda x: x['original_name'])
            
            first_name = lot_data_list[0]['original_name']
            prefix = first_name.split('-')[0] if '-' in first_name else "1"
            
            for idx, lot_data in enumerate(lot_data_list, start=1):
                new_name = f"{prefix}-{idx:02d}"
                lot_data['lot'].write({'name': new_name})

        self.picking_id.write({'worksheet_imported': True})

        message = f'✓ Se actualizaron {lines_updated} lotes con medidas reales.'
        if total_missing_pieces > 0:
            message += f'\n⚠️ MATERIAL FALTANTE:\n• Piezas eliminadas: {total_missing_pieces}\n• Total m² reducidos: {total_missing_m2:.2f} m²'

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Worksheet Procesado Correctamente',
                'message': message,
                'type': 'warning' if total_missing_pieces > 0 else 'success',
                'sticky': True if total_missing_pieces > 0 else False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }

    def _get_data_from_spreadsheet(self):
        pl_wizard = self.env['packing.list.import.wizard'].create({'picking_id': self.picking_id.id})
        doc = self.ws_spreadsheet_id 
        
        data = pl_wizard._load_spreadsheet_json(doc)
        if not data: return []

        revisions = self.env['spreadsheet.revision'].sudo().with_context(active_test=False).search([
            ('res_model', '=', 'documents.document'), ('res_id', '=', doc.id)
        ], order='id asc')

        from .packing_list_import_wizard import _PLCellsIndex
        
        all_rows = []
        for sheet in data.get('sheets', []):
            idx = _PLCellsIndex()
            idx.ingest_cells(sheet.get('cells', {}))
            
            for rev in revisions:
                try:
                    cmds = json.loads(rev.commands) if isinstance(rev.commands, str) else rev.commands
                    if isinstance(cmds, dict) and cmds.get('type') == 'REMOTE_REVISION':
                        idx.apply_revision_commands(cmds.get('commands', []), sheet.get('id'))
                except: continue
            
            product = pl_wizard._resolve_product_from_sheet_id(sheet) or pl_wizard._identify_product_from_sheet(idx)
            if not product: continue
            is_placa = self.picking_id._ws_product_is_placa(product)

            for r in range(3, 250):
                lot_name = str(idx.value(0, r) or '').strip()
                if not lot_name or lot_name == 'Nº Lote': continue

                if is_placa:
                    # Col N (13) = LARGO REAL, Col O (14) = ALTO REAL
                    ancho_r = self._to_float(idx.value(13, r))
                    alto_r = self._to_float(idx.value(14, r))
                    _logger.info(
                        "[WS_IMPORT] placa | fila %s | lote %s | N/Largo(raw)=%r O/Alto(raw)=%r -> alto=%s ancho=%s",
                        r, lot_name, idx.value(13, r), idx.value(14, r), alto_r, ancho_r,
                    )
                    all_rows.append({
                        'product': product,
                        'lot_name': lot_name,
                        'is_placa': True,
                        'alto_real': alto_r,
                        'ancho_real': ancho_r,
                    })
                else:
                    # Formatos: col O (14) = CANT. REAL.
                    qty_r = self._to_float(idx.value(14, r))
                    _logger.info(
                        "[WS_IMPORT] formato | fila %s | lote %s | O(raw)=%r -> qty_real=%s",
                        r, lot_name, idx.value(14, r), qty_r,
                    )
                    all_rows.append({
                        'product': product,
                        'lot_name': lot_name,
                        'is_placa': False,
                        'qty_real': qty_r,
                    })
                    
        return all_rows

    def _get_data_from_excel(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError('Instale openpyxl')
            
        wb = load_workbook(io.BytesIO(base64.b64decode(self.excel_file)), data_only=True)
        all_rows = []
        
        for sheet in wb.worksheets:
            p_info = sheet['B1'].value
            if not p_info: continue
            
            p_code = str(p_info).split('(')[1].split(')')[0].strip() if '(' in str(p_info) else ''
            product = self.env['product.product'].search([
                '|', ('default_code', '=', p_code), ('name', '=', str(p_info).split('(')[0].strip())
            ], limit=1)
            
            if not product: continue
            is_placa = self.picking_id._ws_product_is_placa(product)

            for r in range(4, sheet.max_row + 1):
                lot_name = str(sheet.cell(r, 1).value or '').strip()
                if not lot_name: continue

                if is_placa:
                    # Plantilla WS: col 15 = Largo Real, col 16 = Alto Real
                    all_rows.append({
                        'product': product,
                        'lot_name': lot_name,
                        'is_placa': True,
                        'ancho_real': self._to_float(sheet.cell(r, 15).value),
                        'alto_real': self._to_float(sheet.cell(r, 16).value),
                    })
                else:
                    # Formatos: col 15 = CANT. REAL.
                    all_rows.append({
                        'product': product,
                        'lot_name': lot_name,
                        'is_placa': False,
                        'qty_real': self._to_float(sheet.cell(r, 15).value),
                    })
        return all_rows

    def _to_float(self, val):
        if val is None or val == '': return 0.0
        try:
            return float(str(val).replace(',', '.'))
        except:
            return 0.0