# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import io
import base64
import logging
import json
import re

_logger = logging.getLogger(__name__)

class StockPicking(models.Model):
    _inherit = 'stock.picking'
    
    packing_list_file = fields.Binary(string='Packing List (Archivo)', attachment=True, copy=False)
    packing_list_filename = fields.Char(string='Nombre del archivo', copy=False)
    spreadsheet_id = fields.Many2one('documents.document', string='Spreadsheet Packing List', copy=False)
    has_packing_list = fields.Boolean(string='Tiene Packing List', compute='_compute_has_packing_list', store=True)
    packing_list_imported = fields.Boolean(string='Packing List Importado', default=False, copy=False)
    
    ws_spreadsheet_id = fields.Many2one('documents.document', string='Spreadsheet Worksheet', copy=False)
    worksheet_file = fields.Binary(string='Worksheet Exportado', attachment=True, copy=False)
    worksheet_filename = fields.Char(string='Nombre del Worksheet', copy=False)
    worksheet_imported = fields.Boolean(string='Worksheet Importado', default=False, copy=False)

    supplier_access_ids = fields.One2many('stock.picking.supplier.access', 'picking_id', string="Links Proveedor")
    
    @api.depends('packing_list_file', 'spreadsheet_id', 'supplier_access_ids')
    def _compute_has_packing_list(self):
        for rec in self:
            rec.has_packing_list = bool(rec.packing_list_file or rec.spreadsheet_id or rec.supplier_access_ids)

    # -------------------------------------------------------------------------
    #  LOGICA DE SPREADSHEET (Lectura/Escritura Portal)
    # -------------------------------------------------------------------------

    def get_packing_list_data_for_portal(self):
        """
        Lee el Spreadsheet actual y devuelve la lista de filas
        formateada para el JS del Portal.
        """
        self.ensure_one()
        rows = []
        
        if not self.spreadsheet_id or not self.spreadsheet_id.spreadsheet_data:
            return rows

        try:
            # Intentar leer data cruda
            raw_data = self.spreadsheet_id.spreadsheet_data
            data = json.loads(raw_data.decode('utf-8') if isinstance(raw_data, bytes) else raw_data)
        except Exception as e:
            _logger.warning(f"Error leyendo JSON del spreadsheet: {e}")
            return rows

        # Mapa de Hojas -> Productos
        sheets = data.get('sheets', [])
        
        for sheet in sheets:
            cells = sheet.get('cells', {})
            # B1 contiene el nombre del producto según nuestra plantilla
            b1_val = cells.get("B1", {}).get("content", "")
            
            if not b1_val:
                continue

            # Buscar producto basado en el header
            p_ref = str(b1_val).split('(')[0].strip()
            product = self.env['product.product'].search([
                '|', ('name', 'ilike', p_ref), ('default_code', 'ilike', p_ref)
            ], limit=1)
            
            if not product:
                continue

            # Iterar filas desde la 4 (indice 3) hasta encontrar vacio
            # Estructura: A=Grosor, B=Alto, C=Ancho, D=Color, E=Bloque, G=Tipo, J=Contenedor
            row_idx = 3
            while True:
                idx_str = str(row_idx + 1)
                
                # Si no hay Alto (Col B), asumimos fin de lista o fila vacía
                b_cell = cells.get(f"B{idx_str}", {})
                if not b_cell or not b_cell.get("content"):
                    # Check de seguridad: si tampoco hay ancho (Col C), asumimos fin
                    if not cells.get(f"C{idx_str}", {}).get("content"):
                        # Revisar 3 filas más por si dejaron huecos vacíos
                        found_next = False
                        for lookahead in range(1, 4):
                            if cells.get(f"B{row_idx + 1 + lookahead}", {}).get("content"):
                                found_next = True
                                break
                        if not found_next:
                            break
                        else:
                            row_idx += 1
                            continue

                # Helper para extraer float o string
                def get_val(col, type_cast=str):
                    val = cells.get(f"{col}{idx_str}", {}).get("content", "")
                    if type_cast == float:
                        try: 
                            val_str = str(val).replace(',', '.')
                            return float(val_str)
                        except: 
                            return 0.0
                    return str(val).strip()

                alto = get_val("B", float)
                ancho = get_val("C", float)

                if alto > 0 and ancho > 0:
                    rows.append({
                        'product_id': product.id,
                        'grosor': get_val("A", float),
                        'alto': alto,
                        'ancho': ancho,
                        'color': get_val("D"),
                        'bloque': get_val("E"),
                        'contenedor': get_val("J"),
                        'tipo': get_val("G") or 'placa'
                    })
                
                row_idx += 1
                if row_idx > 2000: break # Limite de seguridad

        return rows

    def update_packing_list_from_portal(self, rows):
        """
        Toma los datos crudos del portal y los escribe en el Spreadsheet.
        Si el Spreadsheet no existe, lo crea primero.
        """
        self.ensure_one()
        
        # 1. Asegurar que existe el Spreadsheet
        if not self.spreadsheet_id:
            self.action_open_packing_list_spreadsheet()
        
        doc = self.spreadsheet_id
        if not doc.spreadsheet_data:
            raise UserError("El archivo de hoja de cálculo está corrupto o vacío.")

        # 2. Cargar datos JSON del Spreadsheet
        try:
            raw_data = doc.spreadsheet_data
            data = json.loads(raw_data.decode('utf-8') if isinstance(raw_data, bytes) else raw_data)
        except Exception as e:
            raise UserError(f"Error al leer el Spreadsheet: {e}")

        # 3. Mapear Productos a Hojas (Sheets)
        # Buscamos en la celda B1 de cada hoja para ver qué producto es.
        product_sheet_map = {} # { product_id: sheet_object }
        
        sheets = data.get('sheets', [])
        for sheet in sheets:
            cells = sheet.get('cells', {})
            # B1 es col 1, row 0 -> clave "B1" o indexada
            # Odoo guarda keys como "B1", "A10", etc.
            b1_val = cells.get("B1", {}).get("content", "")
            
            if b1_val:
                # El formato es "Nombre (Codigo)" o similar. Buscamos coincidencia.
                # Extraemos código o nombre.
                p_ref = str(b1_val).split('(')[0].strip()
                product = self.env['product.product'].search([
                    '|', ('name', 'ilike', p_ref), ('default_code', 'ilike', p_ref)
                ], limit=1)
                
                if product:
                    product_sheet_map[product.id] = sheet
                    
                    # --- LIMPIEZA DE DATOS PREVIOS ---
                    # Eliminamos celdas de datos (Filas > 3) para reescribir limpio
                    # y evitar duplicados o datos viejos si se borraron filas en el portal.
                    keys_to_remove = []
                    for key in list(cells.keys()):
                        # Identificar si la celda está en una fila de datos (A4, B4, C4...)
                        # Regex busca letra + numero
                        match = re.match(r'^([A-Z]+)(\d+)$', key)
                        if match:
                            row_num = int(match.group(2))
                            # Fila 1 (Header), Fila 2 (Info), Fila 3 (Titulos) -> Datos empiezan en Fila 4
                            if row_num >= 4:
                                keys_to_remove.append(key)
                    
                    for k in keys_to_remove:
                        del cells[k]

        # 4. Escribir filas
        # Columnas según action_open_packing_list_spreadsheet:
        # A=Grosor, B=Alto, C=Ancho, D=Color, E=Bloque, F=Atado, G=Tipo, H=Grupo, I=Pedimento, J=Contenedor, K=RefProv, L=Notas
        
        # Agrupar filas por producto para controlar indices de inserción
        rows_by_product = {}
        for row in rows:
            try:
                pid = int(row.get('product_id'))
                if pid not in rows_by_product:
                    rows_by_product[pid] = []
                rows_by_product[pid].append(row)
            except: continue

        # Procesar escritura
        for pid, prod_rows in rows_by_product.items():
            sheet = product_sheet_map.get(pid)
            if not sheet:
                _logger.warning(f"No se encontró hoja para producto ID {pid}")
                continue

            # Empezamos a escribir en la fila 4 (indice 3 para logica 0-based, pero aqui usamos claves excel 1-based)
            current_row = 4
            
            for row in prod_rows:
                # Helper para escribir
                def set_c(col_letter, val):
                    if val is not None:
                        if 'cells' not in sheet: sheet['cells'] = {}
                        sheet['cells'][f"{col_letter}{current_row}"] = {"content": str(val)}

                set_c("A", row.get('grosor', ''))
                set_c("B", row.get('alto', ''))
                set_c("C", row.get('ancho', ''))
                set_c("D", row.get('color', ''))
                set_c("E", row.get('bloque', ''))
                # F (Atado) - No viene del portal simple, dejar vacio o default
                set_c("G", row.get('tipo', 'placa')) # Tipo
                # H, I - Vacios
                set_c("J", row.get('contenedor', ''))
                # K (Ref Prov) - Opcional
                set_c("L", "Actualizado Portal")
                
                current_row += 1

        # 5. Guardar cambios en el documento
        new_json = json.dumps(data)
        doc.write({
            'spreadsheet_data': new_json,
            # Importante: Limpiar snapshot para forzar que se vea lo nuevo
            'spreadsheet_snapshot': False, 
        })
        
        # Eliminar revisiones antiguas para evitar conflictos de sincronización
        self.env['spreadsheet.revision'].sudo().search([
            ('res_model', '=', 'documents.document'),
            ('res_id', '=', doc.id)
        ]).unlink()

        return True

    # -------------------------------------------------------------------------
    # FUNCIONALIDAD ORIGINAL DE PROCESAMIENTO (IMPORT WIZARD)
    # -------------------------------------------------------------------------

    def process_external_pl_data(self, json_data):
        """ 
        Legacy: Mantenido por compatibilidad si se llamara externamente.
        Ahora se prefiere usar el wizard 'packing.list.import.wizard' que lee el spreadsheet.
        """
        return True

    # -------------------------------------------------------------------------
    # UTILS SPREADSHEET
    # -------------------------------------------------------------------------

    def _format_cell_val(self, val):
        if val is None or val is False:
            return ""
        if isinstance(val, (int, float)):
            return str(val)
        result = str(val).strip()
        return result if result else ""

    def _make_cell(self, val, style=None):
        content = self._format_cell_val(val)
        cell = {"content": content}
        if style is not None:
            cell["style"] = style
        return cell

    def _get_col_letter(self, n):
        string = ""
        while n >= 0:
            n, remainder = divmod(n, 26)
            string = chr(65 + remainder) + string
            n -= 1
        return string

    # -------------------------------------------------------------------------
    # GESTIÓN DE PACKING LIST INTERNO (ETAPA 1)
    # -------------------------------------------------------------------------
    
    def action_open_packing_list_spreadsheet(self):
        """ Crea o abre el Spreadsheet para el Packing List inicial. """
        self.ensure_one()
        
        if self.picking_type_code != 'incoming':
            raise UserError('Esta acción solo está disponible para Recepciones (Entradas).')
        
        if self.worksheet_imported:
            raise UserError('El Worksheet ya fue procesado. No es posible modificar el Packing List.')
            
        if not self.spreadsheet_id:
            products = self.move_ids.mapped('product_id')
            if not products:
                raise UserError('No hay productos cargados en esta operación.')

            folder = self.env['documents.document'].search([('type', '=', 'folder')], limit=1)
            headers = [
                'Grosor (cm)', 'Alto (m)', 'Ancho (m)', 'Color', 'Bloque', 'Atado', 
                'Tipo', 'Grupo', 'Pedimento', 'Contenedor', 'Ref. Proveedor', 'Notas'
            ]
            
            sheets = []
            for index, product in enumerate(products):
                cells = {}
                cells["A1"] = self._make_cell("PRODUCTO:")
                p_name = self._format_cell_val(product.name)
                p_code = self._format_cell_val(product.default_code)
                cells["B1"] = self._make_cell(f"{p_name} ({p_code})")
                
                for i, header in enumerate(headers):
                    col_letter = self._get_col_letter(i)
                    cells[f"{col_letter}3"] = self._make_cell(header, style=1)

                sheet_name = (product.default_code or product.name)[:31]
                # Evitar duplicados
                dedup_idx = 1
                orig_name = sheet_name
                while any(s['name'] == sheet_name for s in sheets):
                    sheet_name = f"{orig_name[:25]}_{dedup_idx}"
                    dedup_idx += 1

                sheets.append({
                    "id": f"pl_sheet_{product.id}",
                    "name": sheet_name,
                    "cells": cells,
                    "colNumber": 12,
                    "rowNumber": 250,
                    "isProtected": True,
                    "protectedRanges": [{"range": "A4:L250", "isProtected": False}]
                })

            spreadsheet_data = {
                "version": 16,
                "sheets": sheets,
                "styles": {
                    "1": {"bold": True, "fillColor": "#366092", "textColor": "#FFFFFF", "align": "center"}
                }
            }

            vals = {
                'name': f'PL: {self.name}.osheet',
                'type': 'binary', 
                'handler': 'spreadsheet',
                'mimetype': 'application/o-spreadsheet',
                'spreadsheet_data': json.dumps(spreadsheet_data, ensure_ascii=False, default=str),
                'res_model': 'stock.picking',
                'res_id': self.id,
            }
            if folder:
                vals['folder_id'] = folder.id

            self.spreadsheet_id = self.env['documents.document'].create(vals)

        return self._action_launch_spreadsheet(self.spreadsheet_id)

    # -------------------------------------------------------------------------
    # GESTIÓN DE WORKSHEET (ETAPA 2)
    # -------------------------------------------------------------------------

    def action_open_worksheet_spreadsheet(self):
        """ Crea un Spreadsheet independiente para el Worksheet con datos bloqueados. """
        self.ensure_one()
        if not self.packing_list_imported:
            raise UserError('Debe procesar primero el Packing List para generar el Worksheet.')

        if not self.ws_spreadsheet_id:
            products = self.move_line_ids.mapped('product_id')
            folder = self.env['documents.document'].search([('type', '=', 'folder')], limit=1)

            headers = [
                'Nº Lote', 'Grosor', 'Alto Teo.', 'Ancho Teo.', 'Color', 'Bloque', 
                'Atado', 'Tipo', 'Grupo', 'Pedimento', 'Contenedor', 'Ref. Prov.', 
                'ALTO REAL (m)', 'ANCHO REAL (m)'
            ]
            
            sheets = []
            for product in products:
                cells = {}
                cells["A1"] = self._make_cell("PRODUCTO:")
                p_name = self._format_cell_val(product.name)
                p_code = self._format_cell_val(product.default_code)
                cells["B1"] = self._make_cell(f"{p_name} ({p_code})")
                
                for i, header in enumerate(headers):
                    col_letter = self._get_col_letter(i)
                    cells[f"{col_letter}3"] = self._make_cell(header, style=2)

                move_lines = self.move_line_ids.filtered(lambda ml: ml.product_id == product and ml.lot_id)
                row_idx = 4
                for ml in move_lines:
                    lot = ml.lot_id
                    cells[f"A{row_idx}"] = self._make_cell(lot.name)
                    cells[f"B{row_idx}"] = self._make_cell(lot.x_grosor)
                    cells[f"C{row_idx}"] = self._make_cell(lot.x_alto)
                    cells[f"D{row_idx}"] = self._make_cell(lot.x_ancho)
                    cells[f"E{row_idx}"] = self._make_cell(lot.x_color)
                    cells[f"F{row_idx}"] = self._make_cell(lot.x_bloque)
                    cells[f"G{row_idx}"] = self._make_cell(lot.x_atado)
                    cells[f"H{row_idx}"] = self._make_cell(lot.x_tipo)
                    cells[f"I{row_idx}"] = self._make_cell(", ".join(lot.x_grupo.mapped('name')) if lot.x_grupo else "")
                    cells[f"J{row_idx}"] = self._make_cell(lot.x_pedimento)
                    cells[f"K{row_idx}"] = self._make_cell(lot.x_contenedor)
                    cells[f"L{row_idx}"] = self._make_cell(lot.x_referencia_proveedor)
                    row_idx += 1

                sheet_name = (product.default_code or product.name)[:31]
                sheets.append({
                    "id": f"ws_sheet_{product.id}",
                    "name": sheet_name,
                    "cells": cells,
                    "colNumber": 14,
                    "rowNumber": max(row_idx + 20, 100),
                    "isProtected": True,
                    "protectedRanges": [{"range": f"M4:N{row_idx + 100}", "isProtected": False}]
                })

            spreadsheet_data = {
                "version": 16,
                "sheets": sheets,
                "styles": {
                    "2": {"bold": True, "fillColor": "#1f5b13", "textColor": "#FFFFFF", "align": "center"}
                }
            }

            vals = {
                'name': f'WS: {self.name}.osheet',
                'type': 'binary', 
                'handler': 'spreadsheet',
                'mimetype': 'application/o-spreadsheet',
                'spreadsheet_data': json.dumps(spreadsheet_data, ensure_ascii=False, default=str),
                'res_model': 'stock.picking',
                'res_id': self.id,
            }
            if folder:
                vals['folder_id'] = folder.id

            self.ws_spreadsheet_id = self.env['documents.document'].create(vals)

        return self._action_launch_spreadsheet(self.ws_spreadsheet_id)

    # -------------------------------------------------------------------------
    # FUNCIONES DE APOYO Y EXPORTACIÓN EXCEL
    # -------------------------------------------------------------------------

    def _action_launch_spreadsheet(self, doc):
        """ Dispara la apertura del documento. """
        doc_sudo = doc.sudo()
        for method in ["action_open_spreadsheet", "action_open", "access_content"]:
            if hasattr(doc_sudo, method):
                try:
                    action = getattr(doc_sudo, method)()
                    if action: return action
                except: continue
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'documents.document',
            'res_id': doc.id,
            'view_mode': 'form',
            'target': 'current',
            'context': {'request_handler': 'spreadsheet'}
        }

    def action_download_packing_template(self):
        """ Descarga Excel para el Packing List. """
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side
        except ImportError:
            raise UserError('Instale openpyxl')
            
        wb = Workbook(); wb.remove(wb.active)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for product in self.move_ids.mapped('product_id'):
            ws = wb.create_sheet(title=(product.default_code or product.name)[:31])
            ws['A1'] = 'PRODUCTO:'; ws['B1'] = f'{product.name} ({product.default_code or ""})'
            headers = ['Grosor (cm)', 'Alto (m)', 'Ancho (m)', 'Color', 'Bloque', 'Atado', 'Tipo', 'Grupo', 'Pedimento', 'Contenedor', 'Ref. Proveedor', 'Notas']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num); cell.value = header; cell.fill = header_fill; cell.font = header_font; cell.border = border
            for row in range(4, 54):
                for col in range(1, 13): ws.cell(row=row, column=col).border = border
                
        output = io.BytesIO(); wb.save(output)
        filename = f'Plantilla_PL_{self.name}.xlsx'
        self.write({'packing_list_file': base64.b64encode(output.getvalue()), 'packing_list_filename': filename})
        return {'type': 'ir.actions.act_url', 'url': f'/web/content?model=stock.picking&id={self.id}&field=packing_list_file&filename={filename}&download=true', 'target': 'self'}

    def action_download_worksheet(self):
        """ Descarga Excel para el Worksheet. """
        self.ensure_one()
        if not self.packing_list_imported: raise UserError('Importe primero el Packing List.')
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side
        except ImportError: raise UserError('Instale openpyxl')
        
        wb = Workbook(); wb.remove(wb.active)
        header_fill = PatternFill(start_color='1f5b13', end_color='1f5b13', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        data_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        editable_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for product in self.move_line_ids.mapped('product_id'):
            ws = wb.create_sheet(title=(product.default_code or product.name)[:31])
            ws['A1'] = 'PRODUCTO:'; ws['B1'] = f'{product.name} ({product.default_code or ""})'
            headers = ['Lote', 'Grosor', 'Alto Teo.', 'Ancho Teo.', 'Color', 'Bloque', 'Atado', 'Tipo', 'Grupo', 'Pedimento', 'Contenedor', 'Ref. Prov', 'Cantidad', 'Alto Real', 'Ancho Real']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num); cell.value = header; cell.fill = header_fill; cell.font = header_font; cell.border = border
            
            curr = 4
            for ml in self.move_line_ids.filtered(lambda x: x.product_id == product):
                ws.cell(row=curr, column=1, value=ml.lot_id.name).fill = data_fill
                ws.cell(row=curr, column=2, value=ml.lot_id.x_grosor).fill = data_fill
                ws.cell(row=curr, column=3, value=ml.lot_id.x_alto).fill = data_fill
                ws.cell(row=curr, column=4, value=ml.lot_id.x_ancho).fill = data_fill
                ws.cell(row=curr, column=13, value=ml.qty_done).fill = data_fill
                for col in range(1, 14): ws.cell(row=curr, column=col).border = border
                ws.cell(row=curr, column=14).fill = editable_fill; ws.cell(row=curr, column=14).border = border
                ws.cell(row=curr, column=15).fill = editable_fill; ws.cell(row=curr, column=15).border = border
                curr += 1
                
        output = io.BytesIO(); wb.save(output)
        filename = f'Worksheet_{self.name}.xlsx'
        self.write({'worksheet_file': base64.b64encode(output.getvalue()), 'worksheet_filename': filename})
        return {'type': 'ir.actions.act_url', 'url': f'/web/content?model=stock.picking&id={self.id}&field=worksheet_file&filename={filename}&download=true', 'target': 'self'}

    # -------------------------------------------------------------------------
    # ACCIONES DE WIZARDS
    # -------------------------------------------------------------------------

    def action_import_packing_list(self):
        self.ensure_one()
        
        if self.worksheet_imported:
            raise UserError('El Worksheet ya fue procesado. No es posible reprocesar el Packing List.')
        
        title = 'Aplicar Cambios al PL' if self.packing_list_imported else 'Importar Packing List'
        return {
            'name': title, 
            'type': 'ir.actions.act_window', 
            'res_model': 'packing.list.import.wizard', 
            'view_mode': 'form', 
            'target': 'new', 
            'context': {'default_picking_id': self.id}
        }
    
    def action_import_worksheet(self):
        self.ensure_one()
        return {
            'name': 'Procesar Worksheet (Medidas Reales)', 
            'type': 'ir.actions.act_window', 
            'res_model': 'worksheet.import.wizard', 
            'view_mode': 'form', 
            'target': 'new', 
            'context': {'default_picking_id': self.id}
        }